from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from pathlib import Path


class ExcelReportGenerator:

    @staticmethod
    def write_test_executions_to_excel(
            default_path,
            sheet_names,
            flags,
            releases,
            account_header,
            module_name,
            test_executions):

        tc_ids = []
        results = []

        try:

            for execution in test_executions:
                tc_ids.append(
                    execution.short_description
                )

                results.append(
                    "PASS"
                    if execution.status == "PASS"
                    else "FAIL"
                )

        except AttributeError as e:

            print(
                f"Missing attribute in TestExecution object: {str(e)}"
            )

        except Exception as e:

            print(
                f"Error while processing test executions: {str(e)}"
            )

        ExcelReportGenerator.write_to_excel(
            default_path=default_path,
            sheet_names=sheet_names,
            flags=flags,
            test_case_ids=",".join(tc_ids),
            releases=releases,
            results=",".join(results),
            account_header=account_header,
            module_name=module_name
        )

    @staticmethod
    def write_to_excel(
            default_path,
            sheet_names,
            flags,
            test_case_ids,
            releases,
            results,
            account_header,
            module_name):

        sheet_arr = sheet_names.split(",")
        flag_arr = flags.split(",")
        tc_id_arr = test_case_ids.split(",")
        release_arr = releases.split(",")
        result_arr = results.split(",")

        today = datetime.now().strftime("%Y-%m-%d")

        for s in range(len(sheet_arr)):

            if flag_arr[s].lower() != "yes":
                continue

            sheet_name = sheet_arr[s].strip()

            file_path = Path(default_path) / f"{sheet_name}.xlsx"

            # Load workbook or create new
            if file_path.exists():
                workbook = load_workbook(file_path)
            else:
                workbook = Workbook()

            # Get or create sheet
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)

            # Remove default sheet if empty
            if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
                std = workbook["Sheet"]
                workbook.remove(std)

            # Header row
            header_row = 1

            ExcelReportGenerator.create_header_if_absent(
                sheet,
                header_row,
                1,
                "Module Name"
            )

            ExcelReportGenerator.create_header_if_absent(
                sheet,
                header_row,
                2,
                "Test Case ID"
            )

            result_col_index = ExcelReportGenerator.get_result_column_index(
                sheet=sheet,
                header_row=header_row,
                sheet_name=sheet_name,
                today=today,
                release=release_arr[0],
                account_header=account_header
            )

            # Result styles
            pass_fill = PatternFill(
                start_color="006100",
                end_color="006100",
                fill_type="solid"
            )

            fail_fill = PatternFill(
                start_color="9C0006",
                end_color="9C0006",
                fill_type="solid"
            )

            white_font = Font(color="FFFFFF")

            # Write test results
            for i in range(len(tc_id_arr)):

                tc_id = tc_id_arr[i].strip()
                res = result_arr[i].strip()

                updated = False

                for r in range(2, sheet.max_row + 1):

                    module_cell = sheet.cell(r, 1).value
                    tc_cell = sheet.cell(r, 2).value

                    if module_cell == module_name and tc_cell == tc_id:

                        ExcelReportGenerator.update_result_cell(
                            sheet,
                            r,
                            result_col_index,
                            res,
                            pass_fill,
                            fail_fill,
                            white_font
                        )

                        updated = True
                        break

                if not updated:

                    new_row = sheet.max_row + 1

                    sheet.cell(new_row, 1).value = module_name
                    sheet.cell(new_row, 2).value = tc_id

                    ExcelReportGenerator.update_result_cell(
                        sheet,
                        new_row,
                        result_col_index,
                        res,
                        pass_fill,
                        fail_fill,
                        white_font
                    )

            workbook.save(file_path)

            print(f"Workbook for {sheet_name} updated successfully.")

    @staticmethod
    def write_automation_created_data(
            created_data_name,
            created_date,
            created_time):

        sheet_name = "AutomationData"
        file_path = Path(f"{sheet_name}.xlsx")

        if file_path.exists():
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Remove default sheet
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            std = workbook["Sheet"]
            workbook.remove(std)

        headers = ["Name", "Date", "Time"]

        header_font = Font(
            bold=True,
            color="000000"
        )

        header_fill = PatternFill(
            start_color="B4C6E7",
            end_color="B4C6E7",
            fill_type="solid"
        )

        # Create headers
        for col, header in enumerate(headers, start=1):

            cell = sheet.cell(1, col)

            if not cell.value:
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill

        # Add data row
        new_row = sheet.max_row + 1

        sheet.cell(new_row, 1).value = created_data_name
        sheet.cell(new_row, 2).value = created_date
        sheet.cell(new_row, 3).value = created_time

        workbook.save(file_path)

        print("Automation data written successfully.")

    @staticmethod
    def create_header_if_absent(
            sheet,
            row,
            column,
            value):

        cell = sheet.cell(row, column)

        if not cell.value:

            cell.value = value

            cell.font = Font(
                bold=True,
                color="000000"
            )

            cell.fill = PatternFill(
                start_color="B4C6E7",
                end_color="B4C6E7",
                fill_type="solid"
            )

    @staticmethod
    def get_result_column_index(
            sheet,
            header_row,
            sheet_name,
            today,
            release,
            account_header):

        if sheet_name.lower() == "daily":
            dynamic_header = today

        elif sheet_name.lower() == "release":
            dynamic_header = f"Release - {release}"

        else:
            dynamic_header = account_header

        for col in range(3, sheet.max_column + 1):

            cell_value = sheet.cell(header_row, col).value

            if cell_value == dynamic_header:
                return col

        new_col = sheet.max_column + 1

        cell = sheet.cell(header_row, new_col)

        cell.value = dynamic_header

        cell.font = Font(
            bold=True,
            color="000000"
        )

        cell.fill = PatternFill(
            start_color="B4C6E7",
            end_color="B4C6E7",
            fill_type="solid"
        )

        return new_col

    @staticmethod
    def update_result_cell(
            sheet,
            row,
            col,
            result,
            pass_fill,
            fail_fill,
            white_font):

        cell = sheet.cell(row, col)

        cell.value = result

        cell.font = white_font

        if "PASS" in result:
            cell.fill = pass_fill
        else:
            cell.fill = fail_fill

